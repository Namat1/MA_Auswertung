import streamlit as st
import pandas as pd
import datetime
from io import BytesIO
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

st.title("Fahrerauswertung - Einzeln")

uploaded_files = st.file_uploader("Excel-Dateien hochladen", type=["xlsx"], accept_multiple_files=True)

wochentage_deutsch = {
    "Monday": "Montag", "Tuesday": "Dienstag", "Wednesday": "Mittwoch",
    "Thursday": "Donnerstag", "Friday": "Freitag",
    "Saturday": "Samstag", "Sunday": "Sonntag"
}

thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

def get_kw_and_year_sunday_start(datum):
    try:
        dt = pd.to_datetime(datum)
        kw = int(dt.strftime("%U")) + 1
        jahr = dt.year
        if dt.month == 1 and kw >= 52:
            jahr -= 1
        return kw, jahr
    except:
        return None, None

def format_uhrzeit(val):
    try:
        if pd.isna(val):
            return "n. A."
        if isinstance(val, str):
            val = val.strip()
            if val in ["0:00", "00:00", "00:00:00"]:
                return "00:00"
            if ":" in val:
                parts = val.split(":")
                if len(parts) >= 2 and all(p.isdigit() for p in parts[:2]):
                    return f"{int(parts[0]):02d}:{int(parts[1]):02d}"
        elif isinstance(val, (float, int)):
            if val == 0:
                return "00:00"
            stunden = int(val * 24)
            minuten = int((val * 1440) % 60)
            return f"{stunden:02d}:{minuten:02d}"
        elif isinstance(val, (datetime.datetime, pd.Timestamp)):
            return val.strftime("%H:%M")
        elif isinstance(val, datetime.time):
            return val.strftime("%H:%M")
    except:
        return "n. A."
def extract_entries_both_sides(row):
    eintraege = []
    datum = pd.to_datetime(row[14], errors="coerce")
    if pd.isna(datum):
        return []

    kw, jahr = get_kw_and_year_sunday_start(datum)
    wochentag = datum.day_name()
    wochentag_de = wochentage_deutsch.get(wochentag, wochentag)
    datum_formatiert = datum.strftime('%d.%m.%Y')
    datum_komplett = f"{wochentag_de}, {datum_formatiert}"
    uhrzeit = format_uhrzeit(row[8])

    if pd.notna(row[3]) and pd.notna(row[4]):
        name = f"{str(row[3]).strip()} {str(row[4]).strip()}"
        eintraege.append({
            "KW": kw, "Jahr": jahr, "Datum": datum_komplett, "Datum_sortierbar": datum,
            "Name": name, "Tour": row[15], "Uhrzeit": uhrzeit, "LKW": row[11]
        })

    if pd.notna(row[6]) and pd.notna(row[7]):
        name = f"{str(row[6]).strip()} {str(row[7]).strip()}"
        eintraege.append({
            "KW": kw, "Jahr": jahr, "Datum": datum_komplett, "Datum_sortierbar": datum,
            "Name": name, "Tour": row[15], "Uhrzeit": uhrzeit, "LKW": row[11]
        })

    return eintraege

if uploaded_files:
    eintraege_gesamt = []

    for file in uploaded_files:
        try:
            df = pd.read_excel(file, sheet_name="Touren", header=None)
            df = df.iloc[5:].reset_index(drop=True)
            for _, row in df.iterrows():
                eintraege_gesamt.extend(extract_entries_both_sides(row))
        except Exception as e:
            st.error(f"Fehler beim Verarbeiten von {file.name}: {e}")

    if eintraege_gesamt:
        df_final = pd.DataFrame(eintraege_gesamt)

        fahrersuche = st.text_input("Nach Namen suchen (z. B. 'müller')").strip().lower()
        passende_namen = []

        if fahrersuche:
            passende_namen = sorted(
                [name for name in df_final["Name"].dropna().unique()
                 if fahrersuche in name.lower()]
            )

        if passende_namen:
            ausgewaehlter_name = st.selectbox("Passenden Fahrer auswählen", passende_namen)
            df_final = df_final[df_final["Name"] == ausgewaehlter_name]
        else:
            df_final = df_final.iloc[0:0]

        if df_final.empty and fahrersuche and passende_namen:
            st.warning("Für diesen Fahrer wurden keine Touren gefunden.")
        if not df_final.empty:
            df_final.sort_values(by=["Jahr", "KW", "Datum_sortierbar"], inplace=True)

            output = BytesIO()
            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                sheet = "Touren"
                ws = writer.book.create_sheet(title=sheet)
                writer.sheets[sheet] = ws

                start_row = 1
                for (jahr, kw), group in df_final.groupby(["Jahr", "KW"]):
                    group = group.reset_index(drop=True)

                    ws.cell(row=start_row, column=1, value=f"KW {int(kw)} ({int(jahr)})")
                    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=7)
                    cell = ws.cell(row=start_row, column=1)
                    cell.font = Font(bold=True, size=14)
                    cell.alignment = Alignment(horizontal="left")
                    cell.fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
                    start_row += 1

                    header = ["KW", "Jahr", "Datum", "Name", "Tour", "Uhrzeit", "LKW"]
                    for col_num, column_title in enumerate(header, 1):
                        cell = ws.cell(row=start_row, column=col_num, value=column_title)
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                        cell.border = thin_border
                    start_row += 1

                    for row in group.itertuples(index=False):
                        values = [row.KW, row.Jahr, row.Datum, row.Name, row.Tour, row.Uhrzeit, row.LKW]
                        is_samstag = "Samstag" in str(row.Datum)
                        for col_num, value in enumerate(values, 1):
                            cell = ws.cell(row=start_row, column=col_num, value=value)
                            cell.alignment = Alignment(horizontal="left", vertical="center")
                            cell.border = thin_border
                            if is_samstag:
                                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                        start_row += 1

                    start_row += 1
                # Rechte Auswertung mit „davon Samstag“
                summary_labels = ["Tage Krank", "Tage Urlaub", "Tage Arbeit", "Tage Ausgleich"]
                krank_count = df_final["Tour"].astype(str).str.lower().str.contains("krank").sum()
                urlaub_count = df_final["Tour"].astype(str).str.lower().str.contains("urlaub").sum()
                ausgleich_count = df_final["Tour"].astype(str).str.lower().str.contains("ausgleich").sum()
                arbeit_count = df_final["Uhrzeit"].astype(str).apply(lambda x: x.strip() != "n. A.").sum()
                arbeit_samstag_count = df_final[
                    (df_final["Uhrzeit"].astype(str).str.strip() != "n. A.") &
                    (df_final["Datum"].astype(str).str.contains("Samstag"))
                ].shape[0]

                summary_values = [krank_count, urlaub_count, arbeit_count, ausgleich_count]
                summary_col_label = 9
                summary_col_value = 10
                row_cursor = 2

                for idx, label in enumerate(summary_labels):
                    value = summary_values[idx]
                    cell_label = ws.cell(row=row_cursor, column=summary_col_label, value=label)
                    cell_value = ws.cell(row=row_cursor, column=summary_col_value, value=value)
                    cell_label.font = Font(bold=True)
                    cell_label.fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
                    cell_label.alignment = Alignment(horizontal="left", vertical="center")
                    cell_value.alignment = Alignment(horizontal="center", vertical="center")
                    cell_label.border = thin_border
                    cell_value.border = thin_border
                    row_cursor += 1

                    if label == "Tage Arbeit":
                        cell_label = ws.cell(row=row_cursor, column=summary_col_label, value="davon Samstag")
                        cell_value = ws.cell(row=row_cursor, column=summary_col_value, value=arbeit_samstag_count)
                        cell_label.font = Font(bold=True)
                        cell_label.fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
                        cell_label.alignment = Alignment(horizontal="left", vertical="center")
                        cell_value.alignment = Alignment(horizontal="center", vertical="center")
                        cell_label.border = thin_border
                        cell_value.border = thin_border
                        row_cursor += 1

                # Tourenübersicht (gefiltert)
                row_cursor += 2
                ws.cell(row=row_cursor, column=summary_col_label, value="Tourenübersicht:")
                ws.cell(row=row_cursor, column=summary_col_label).font = Font(bold=True, size=12)
                row_cursor += 1

                touren_zaehler = (
                    df_final["Tour"]
                    .dropna()
                    .astype(str)
                    .str.strip()
                    .loc[lambda s: ~s.str.lower().str.contains("krank|urlaub|ausgleich")]
                    .value_counts()
                    .sort_index()
                )

                for tour, count in touren_zaehler.items():
                    cell_name = ws.cell(row=row_cursor, column=summary_col_label, value=tour)
                    cell_count = ws.cell(row=row_cursor, column=summary_col_value, value=f"{count}x")
                    cell_name.alignment = Alignment(horizontal="left", vertical="center")
                    cell_count.alignment = Alignment(horizontal="center", vertical="center")
                    cell_name.font = Font(bold=False)
                    cell_count.font = Font(bold=True)
                    cell_name.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                    cell_count.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                    cell_name.border = thin_border
                    cell_count.border = thin_border
                    row_cursor += 1

                # Auto-Breite
                for col in ws.columns:
                    max_length = 0
                    col_letter = get_column_letter(col[0].column)
                    for cell in col:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    ws.column_dimensions[col_letter].width = int(max_length * 1.5)

            name_sicher = ausgewaehlter_name.replace(" ", "_") if passende_namen else "Touren"
            file_name = f"{name_sicher}_Auswertung.xlsx"

            output.seek(0)
            st.success("Auswertung abgeschlossen.")
            st.download_button("Excel-Datei herunterladen",
                               output,
                               file_name=file_name,
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
